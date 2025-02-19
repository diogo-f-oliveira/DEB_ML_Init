import os
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def format_results_into_dataframe(results_array, col_types, species_list):
    return pd.DataFrame(results_array, columns=col_types['output']['all'], index=species_list)


def create_results_directories_for_dataset(dataset_name, results_path='results'):
    # Create folders to evaluate initialization performance
    deb_model_results_folders = ['parameter_predictions', 'feasibility', 'deb_model_loss', 'initialization']
    for results_type in deb_model_results_folders:
        deb_results_folder = os.path.join(results_path, results_type)
        if not os.path.exists(deb_results_folder):
            os.makedirs(deb_results_folder)

    # Create dataset folder
    base_dataset_folder = os.path.join(results_path, dataset_name)
    if not os.path.exists(base_dataset_folder):
        os.makedirs(base_dataset_folder)

    # Build folders for models
    model_folder = os.path.join(base_dataset_folder, 'models')
    if not os.path.exists(model_folder):
        os.makedirs(model_folder)

    # Build folders for test_performance
    test_performance_folder = os.path.join(base_dataset_folder, 'test_performance')
    if not os.path.exists(test_performance_folder):
        os.makedirs(test_performance_folder)




def fetch_best_results_per_model(results_folder, save=False):
    # Get folders in dataset_folder
    output_cols = [fol for fol in os.listdir(results_folder) if os.path.isdir(os.path.join(results_folder, fol))]
    best_results_df = {}
    for col in output_cols:
        col_results_folder = os.path.join(results_folder, col)
        test_performance_folder = os.path.join(col_results_folder, 'test_performance')
        if not os.path.exists(test_performance_folder):
            continue
        model_results_files = os.listdir(test_performance_folder)
        best_result_per_model_type = {}
        best_results_df[col] = pd.DataFrame()
        for results_file in model_results_files:
            metric, value, model_type = (results_file.split('.')[0]).split('_')
            value = int(value)
            if model_type not in best_result_per_model_type:
                best_result_per_model_type[model_type] = value
                model_results_df = pd.read_csv(os.path.join(col_results_folder, f"test_performance/{results_file}"),
                                               index_col=0)
                model_results_df = pd.DataFrame(data=model_results_df.values, index=[model_type],
                                                columns=model_results_df.columns, dtype=float)
                best_results_df[col] = pd.concat([best_results_df[col], model_results_df])
            elif value > best_result_per_model_type[model_type]:
                best_result_per_model_type[model_type] = value
                model_results_df = pd.read_csv(os.path.join(col_results_folder, f"test_performance/{results_file}"),
                                               index_col=0)
                model_results_df = pd.DataFrame(data=model_results_df.values, index=[model_type],
                                                columns=model_results_df.columns)
                best_results_df[col].loc[model_type] = model_results_df.loc[model_type]
        best_results_df[col].index.name = 'model'
        best_results_df[col].sort_index(inplace=True)
        if save:
            best_results_df[col].to_csv(os.path.join(results_folder, f'{col}__best_models_results.csv'),
                                        float_format='%.6f')

    return best_results_df


def get_best_model_file(results_folder, model_type=None, metric='MAPE'):
    model_files = os.listdir(os.path.join(results_folder, 'models'))
    best_model_file = None
    if metric in ['MAPE', 'MSE', 'GEF', 'sMAPE']:
        best_score = float('inf')
        method = min
    elif metric in ['R2', 'D2']:
        best_score = -float('inf')
        method = max

    for model_file in model_files:
        file_metric, score_str, model_name = (model_file.split('.')[0]).split('_')
        score = float(score_str[:1] + '.' + score_str[1:])
        if file_metric != metric:
            continue
        # Check the model matches the intended model type
        if model_type is None or model_name == model_type:
            # Check if score is better
            if method(score, best_score) == score:
                best_score = score
                best_model_file = model_file
    return best_model_file


def generate_excel_results(results_folder, dataset_name, existing_file=None):
    """Generates an Excel sheet with a table with the regression results for each parameter. Assumes the .csv files
    with the best results per model has been generated already."""

    header_mapping = {
        'model': 'model',
        'r2_score': 'r2_score (↑)',
        'mean_squared_error': 'mse (↓)',
        'max_error': 'max_error (↓)',
        'mean_absolute_percentage_error': 'mape (↓)',
        'mean_gamma_deviance': 'mgd (↓)',
        'd2_score': 'd2_score (↑)',
        'mean_deb_loss': 'deb_loss (↓)',
    }

    # Get .csv files
    csv_files = [path for path in os.listdir(results_folder) if path.endswith('.csv')]

    # Create a new workbook and select the active worksheet
    if existing_file is None:
        wb = Workbook()
        ws = wb.active
        ws.title = dataset_name
        excel_file = f'{results_folder}/best_model_results.xlsx'
    else:
        wb = load_workbook(existing_file)
        ws = wb.create_sheet(title=dataset_name)
        excel_file = existing_file

    # Define a named style for numerical formatting to 5 decimal places
    num_format_name = "num_format"
    if num_format_name not in wb.named_styles:  # Add style only if it doesn't already exist
        num_format = NamedStyle(name=num_format_name, number_format="0.00000")
        wb.add_named_style(num_format)
    # Font style for bold
    bold_font = Font(bold=True)
    # Define border styles
    header_border = Border(
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    bottom_border = Border(
        bottom=Side(style='medium')
    )

    # Initialize starting row
    start_row = 3

    for file in csv_files:
        # Load the best results for the parameter
        par = file.split('__')[0]
        df = pd.read_csv(os.path.join(results_folder, file), index_col=0)

        # Set the column widths
        if start_row == 3:  # Only set widths for the first table to avoid redundant settings
            # Set width for the index column (first column)
            ws.column_dimensions['A'].width = 22
            # Set width for all other columns with values
            for col_idx in range(2, df.shape[1] + 2):  # Adjust range as index is counted
                col_letter = ws.cell(row=start_row, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 13

                # Write parameter name and header
        ws.cell(row=start_row, column=1, value=par)
        start_row += 1
        # Rename columns using the provided header mapping
        df = df.rename(columns=header_mapping)
        # Ensure all numerical columns are formatted to 5 decimal places
        df = df.apply(lambda x: round(x, 5) if pd.api.types.is_numeric_dtype(x) else x)

        # Find the minimum value in each column, excluding the header
        min_values = df.min(numeric_only=True)
        best_values = []
        for col in df.columns:
            if '↓' in col:
                best_values.append(df[col].min())
            else:
                best_values.append(df[col].max())

        # Convert DataFrame to rows and append them to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df.reset_index(), index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Apply the numerical format to all numerical cells (skip header)
                if r_idx > start_row and isinstance(value, (int, float)):
                    cell.style = num_format_name
                    # Check if the current value is the minimum in its column
                    if value == best_values[c_idx - 2]:
                        cell.font = bold_font
                # Add borders to header row
                if r_idx == start_row:
                    cell.border = header_border
                if r_idx == start_row + len(df):
                    cell.border = bottom_border

        # Update start_row for the next table, leaving two empty rows
        start_row = ws.max_row + 3  # Add 3 to leave 2 empty rows between tables

    # Save the workbook to an Excel file
    wb.save(excel_file)
    print(f'Results file saved in {excel_file}')


def copy_debtool_results_files(species_list, destination_folder, copy_pars_init_file=True, copy_results_mat_file=False,
                               all_species_folder=None):
    if all_species_folder is None:
        filepaths = {}
        with open('../../filepaths.csv', 'r') as f:
            for line in f:
                name, path = line.rstrip('\n').split(',')
                filepaths[name] = os.path.abspath(path)
        all_species_folder = os.path.abspath(filepaths['species_folder'])

    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    for species in species_list:
        if copy_pars_init_file:
            source_file = os.path.join(all_species_folder, species, f'pars_init_{species}.m')
            destination_file = os.path.join(destination_folder, f'pars_init_{species}.m')

            if os.path.exists(source_file):
                shutil.copy(source_file, destination_file)
            else:
                print(f"pars_init_{species}.m file not found")

        if copy_results_mat_file:
            source_file = os.path.join(all_species_folder, species, f'results_{species}.mat')
            destination_file = os.path.join(destination_folder, f'results_{species}.mat')

            if os.path.exists(source_file):
                shutil.copy(source_file, destination_file)
            else:
                print(f"results_{species}.mat file not found")


if __name__ == '__main__':
    # dataset_name = 'no_pub_weight'
    # dataset_name = 'ratio_no_pub_weight'
    # dataset_name = 'ratio_output_no_pub_age'
    # # dataset_name = 'bijection_input'
    #
    # dataset_name += '_taxonomy'
    # dataset_name += '_ecocodes'
    #
    # dataset_results_folder = f'results/{dataset_name}'
    #
    # model_file = get_best_model_file(
    #     results_folder=dataset_results_folder,
    #     model_type='MultiTaskElasticNet',
    #     metric='MAPE'
    # )
    # print(model_file)
    # fetch_best_results_per_model(results_folder=dataset_results_folder)
    # generate_excel_results(results_folder=dataset_results_folder,
    #                        dataset_name=dataset_name,
    #                        # existing_file=f'results/Best model comparison.xlsx'
    #                        )

    copy_debtool_results_files(
        species_list=['Caenorhabditis_elegans', 'Caranx_ignobilis', 'Geococcyx_californianus', 'Torpedo_marmorata',
                      'Pleurobrachia_pileus', 'Turdus_merula', 'Gallotia_galloti', 'Diplectrum_formosum',
                      'Octopus_joubini', 'Myiarchus_cinerascens', 'Somateria_mollissima', 'Paranotothenia_magellanica',
                      'Sillago_sihama', 'Stolephorus_waitei', 'Rhombosolea_plebeia', 'Pseudopleuronectes_yokohamae',
                      'Microcondylaea_bonellii', 'Trichiurus_lepturus', 'Aerodramus_fuciphagus', 'Sula_dactylatra',
                      'Scomberomorus_maculatus', 'Macquaria_ambigua', 'Apteryx_mantelli', 'Anchoa_panamensis',
                      'Synodontis_nebulosus', 'Euthynnus_alletteratus', ],
        destination_folder=r"C:\Users\diogo\Downloads\pars_init files in the loss function minimum",
        copy_pars_init_file=True,
        copy_results_mat_file=False,
    )
